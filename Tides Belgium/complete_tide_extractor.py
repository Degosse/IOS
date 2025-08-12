#!/usr/bin/env python3
"""
Complete Belgian Tide Data Extraction System
Extracts full year data from Excel files for all 5 Belgian stations.
Supports both 2025 and 2026 data with proper July/August separation.
"""

import openpyxl
import json
import os
from datetime import datetime

def parse_time(time_val):
    """Parse time from Excel cell value"""
    if time_val is None:
        return None
    if hasattr(time_val, 'hour'):
        return f'{time_val.hour:02d}:{time_val.minute:02d}'
    time_str = str(time_val).strip()
    if ':' in time_str:
        parts = time_str.split(':')
        if len(parts) >= 2:
            try:
                hour = int(parts[0])
                minute = int(parts[1])
                return f'{hour:02d}:{minute:02d}'
            except ValueError:
                pass
    return None

def parse_height(height_val):
    """Parse height from Excel cell value"""
    if height_val is None:
        return None
    try:
        if isinstance(height_val, (int, float)):
            return float(height_val)
        height_str = str(height_val).replace(',', '.')
        return float(height_str)
    except ValueError:
        return None

def get_months_from_sheet_name(sheet_name):
    """Convert Dutch sheet names to month numbers"""
    month_map = {
        'jan-feb': [1, 2],
        'mrt-apr': [3, 4], 
        'mei-jun': [5, 6],
        'jul-aug': [7, 8],
        'sept-okt': [9, 10],
        'nov-dec': [11, 12]
    }
    return month_map.get(sheet_name.lower(), [])

def extract_station_full_year(excel_path, station_name, year):
    """Extract complete year of tide data from Excel file"""
    print(f"\\n📊 Processing {station_name.upper()} for year {year}...")
    
    if not os.path.exists(excel_path):
        print(f"❌ Excel file not found: {excel_path}")
        return []
    
    wb = openpyxl.load_workbook(excel_path)
    all_tides = []
    
    for sheet_name in wb.sheetnames:
        if '-' not in sheet_name.lower():
            continue
            
        months = get_months_from_sheet_name(sheet_name)
        if not months:
            continue
            
        print(f"  📋 Processing sheet: {sheet_name} (months {months})")
        ws = wb[sheet_name]
        
        # Start from row 4 (after headers: NIEUWPOORT, Hoogwater/Laagwater, Datum/uu:mm/m TAW)
        for row in range(4, ws.max_row + 1):
            day_val = ws.cell(row=row, column=1).value
            
            if not isinstance(day_val, (int, float)) or day_val <= 0 or day_val > 31:
                continue
                
            day = int(day_val)
            
            # Extract data for each month in this sheet
            for month_idx, month in enumerate(months):
                try:
                    # Validate date exists
                    datetime(year, month, day)
                    date_str = f'{year}-{month:02d}-{day:02d}'
                    
                    # Column mapping based on month within sheet
                    if month_idx == 0:  # First month (e.g., July in jul-aug)
                        tide_columns = [(3, 4), (5, 6)]  # Time-Height pairs for first month
                    else:  # Second month (e.g., August in jul-aug)
                        tide_columns = [(10, 11)]  # Time-Height pairs for second month
                    
                    # Extract tides from main row
                    for time_col, height_col in tide_columns:
                        time_val = ws.cell(row=row, column=time_col).value
                        height_val = ws.cell(row=row, column=height_col).value
                        
                        time_str = parse_time(time_val)
                        height = parse_height(height_val)
                        
                        if time_str and height is not None:
                            tide_type = 'high' if height >= 2.5 else 'low'
                            all_tides.append({
                                'date': date_str,
                                'time': time_str,
                                'height': round(height, 2),
                                'type': tide_type
                            })
                    
                    # Check continuation row
                    next_row = row + 1
                    if next_row <= ws.max_row:
                        next_day_val = ws.cell(row=next_row, column=1).value
                        if not isinstance(next_day_val, (int, float)):
                            for time_col, height_col in tide_columns:
                                time_val = ws.cell(row=next_row, column=time_col).value
                                height_val = ws.cell(row=next_row, column=height_col).value
                                
                                time_str = parse_time(time_val)
                                height = parse_height(height_val)
                                
                                if time_str and height is not None:
                                    tide_type = 'high' if height >= 2.5 else 'low'
                                    all_tides.append({
                                        'date': date_str,
                                        'time': time_str,
                                        'height': round(height, 2),
                                        'type': tide_type
                                    })
                    
                except ValueError:
                    # Invalid date (e.g., Feb 30), continue
                    continue
    
    # Sort by date and time, remove duplicates
    all_tides.sort(key=lambda x: (x['date'], x['time']))
    
    unique_tides = []
    seen = set()
    for tide in all_tides:
        key = f"{tide['date']}_{tide['time']}_{tide['height']}"
        if key not in seen:
            seen.add(key)
            unique_tides.append(tide)
    
    print(f"  ✅ Extracted {len(unique_tides)} unique tides")
    return unique_tides

def main():
    """Main extraction process for all stations and years"""
    
    # Station configurations
    stations = [
        {'name': 'nieuwpoort', 'file_2025': 'xlsx-getijtabellen-taw-2025/Nieuwpoort2025_mTAW.xlsx', 'file_2026': 'xlsx-getijtabellen-taw-2026/Nieuwpoort_2026_mTAW.xlsx'},
        {'name': 'oostende', 'file_2025': 'xlsx-getijtabellen-taw-2025/Oostende2025_mTAW.xlsx', 'file_2026': 'xlsx-getijtabellen-taw-2026/Oostende_2026_mTAW.xlsx'},
        {'name': 'blankenberge', 'file_2025': 'xlsx-getijtabellen-taw-2025/Blankenberge2025_mTAW.xlsx', 'file_2026': 'xlsx-getijtabellen-taw-2026/Blankenberge_2026_mTAW.xlsx'},
        {'name': 'zeebrugge', 'file_2025': 'xlsx-getijtabellen-taw-2025/Zeebrugge2025_mTAW.xlsx', 'file_2026': 'xlsx-getijtabellen-taw-2026/Zeebrugge_2026_mTAW.xlsx'},
        {'name': 'antwerpen', 'file_2025': 'xlsx-getijtabellen-taw-2025/Antwerpen2025_mTAW.xlsx', 'file_2026': 'xlsx-getijtabellen-taw-2026/Antwerpen_2026_mTAW.xlsx'},
    ]
    
    # Process both years
    for year in [2025, 2026]:
        print(f\"\\n🗓️  PROCESSING YEAR {year}\")
        print(\"=\" * 50)
        
        for station in stations:
            file_key = f'file_{year}'
            if file_key in station:
                excel_path = station[file_key]
                tides = extract_station_full_year(excel_path, station['name'], year)
                
                if tides:
                    output_file = f\"{station['name']}_{year}.json\"
                    with open(output_file, 'w') as f:
                        json.dump(tides, f, indent=2)
                    
                    print(f\"  💾 Saved: {output_file}\")
                    
                    # Show sample of current month data
                    current_month = f\"{year}-08-\"  # August for testing
                    current_tides = [t for t in tides if t['date'].startswith(current_month)]
                    if current_tides:
                        print(f\"  📅 August {year} sample ({len(current_tides)} tides):
                        for tide in current_tides[:4]:  # Show first 4
                            print(f\"    {tide['date']} {tide['time']}: {tide['height']}m ({tide['type']})
                else:
                    print(f\"  ❌ No data extracted for {station['name']} {year}\")
    
    print(f\"\\n🎉 EXTRACTION COMPLETE!\")
    print(\"📱 Deploy these JSON files to your iOS app:\")
    print(\"   • Copy to Xcode project bundle for permanent inclusion\")
    print(\"   • Or copy to iOS Simulator Documents folder for testing\")

if __name__ == \"__main__\":
    main()
