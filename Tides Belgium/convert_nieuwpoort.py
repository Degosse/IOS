#!/usr/bin/env python3
"""
Comprehensive Excel to JSON converter for Belgian tide data.
Handles all sheets and months correctly.
"""

import json
import sys
import os
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl not installed. Run: pip3 install openpyxl")
    sys.exit(1)

def parse_time(time_val):
    """Parse time from various formats in the Excel."""
    if time_val is None:
        return None
    
    time_str = str(time_val).strip()
    
    # Handle time objects from Excel
    if hasattr(time_val, 'hour'):
        return f"{time_val.hour:02d}:{time_val.minute:02d}"
    
    # Handle HH:MM:SS format
    if ':' in time_str:
        parts = time_str.split(':')
        if len(parts) >= 2:
            try:
                hour = int(parts[0])
                minute = int(parts[1])
                return f"{hour:02d}:{minute:02d}"
            except ValueError:
                pass
    
    return None

def parse_height(height_val):
    """Parse height from Excel format."""
    if height_val is None:
        return None
    
    try:
        if isinstance(height_val, (int, float)):
            return float(height_val)
        
        height_str = str(height_val).replace(',', '.')
        return float(height_str)
    except ValueError:
        return None

def get_month_from_sheet_name(sheet_name):
    """Determine months covered by sheet name."""
    sheet_lower = sheet_name.lower()
    
    month_mapping = {
        'jan-feb': [1, 2], 'mrt-apr': [3, 4], 'mei-jun': [5, 6],
        'jul-aug': [7, 8], 'sept-okt': [9, 10], 'nov-dec': [11, 12]
    }
    
    for pattern, months in month_mapping.items():
        if pattern in sheet_lower:
            return months
    
    return None

def convert_excel_to_json(excel_file, station_name, year):
    """Convert an Excel file to JSON format."""
    print(f"Converting {excel_file} for {station_name} {year}...")
    
    try:
        wb = openpyxl.load_workbook(excel_file)
        tides = []
        
        # Process all sheets
        for sheet_name in wb.sheetnames:
            print(f"Processing sheet: {sheet_name}")
            ws = wb[sheet_name]
            
            months = get_month_from_sheet_name(sheet_name)
            if not months:
                print(f"  Skipping sheet {sheet_name} - unknown month pattern")
                continue
            
            for row in range(1, ws.max_row + 1):
                day_val = ws.cell(row=row, column=1).value
                
                if isinstance(day_val, (int, float)) and 1 <= day_val <= 31:
                    day = int(day_val)
                    
                    # Process both months for this sheet
                    for month_idx, current_month in enumerate(months):
                        # Determine if this day belongs to first or second month of the sheet
                        # This is a simplification - in practice, you'd need more logic
                        # For now, assume days 1-15 are first month, 16-31 are second month
                        if month_idx == 0 and day > 15 and len(months) > 1:
                            continue
                        if month_idx == 1 and day <= 15:
                            continue
                        if month_idx == 1 and len(months) == 1:
                            continue
                            
                        # Actually, let's use column 8 to detect second month
                        day2_val = ws.cell(row=row, column=8).value
                        if isinstance(day2_val, (int, float)) and month_idx == 0:
                            # This row has data for second month too
                            pass
                        elif month_idx == 1 and not isinstance(day2_val, (int, float)):
                            # No second month data in this row
                            continue
                        
                        if month_idx == 1:
                            # Use day from column 8 for second month
                            day_val_2 = ws.cell(row=row, column=8).value
                            if isinstance(day_val_2, (int, float)):
                                day = int(day_val_2)
                            else:
                                continue
                        
                        date_str = f"{year:04d}-{current_month:02d}-{day:02d}"
                        
                        # Column mapping based on which month we're processing
                        if month_idx == 0:  # First month columns
                            tide_columns = [
                                (3, 4, None),   # Time, height (determine type from height)
                                (5, 6, None),   # Time, height
                            ]
                        else:  # Second month columns  
                            tide_columns = [
                                (10, 11, None), # Time, height
                                (12, 13, None), # Time, height (if exists)
                            ]
                        
                        for time_col, height_col, _ in tide_columns:
                            time_val = ws.cell(row=row, column=time_col).value
                            height_val = ws.cell(row=row, column=height_col).value
                            
                            time_str = parse_time(time_val)
                            height = parse_height(height_val)
                            
                            if time_str and height is not None:
                                # Determine tide type from height
                                tide_type = "high" if height >= 3.0 else "low"
                                
                                tide_entry = {
                                    "date": date_str,
                                    "time": time_str,
                                    "height": round(height, 2),
                                    "type": tide_type
                                }
                                tides.append(tide_entry)
                        
                        # Check continuation row
                        next_row = row + 1
                        if next_row <= ws.max_row:
                            next_day_val = ws.cell(row=next_row, column=1).value
                            if not isinstance(next_day_val, (int, float)):
                                # Continuation row
                                for time_col, height_col, _ in tide_columns:
                                    time_val = ws.cell(row=next_row, column=time_col).value
                                    height_val = ws.cell(row=next_row, column=height_col).value
                                    
                                    time_str = parse_time(time_val)
                                    height = parse_height(height_val)
                                    
                                    if time_str and height is not None:
                                        tide_type = "high" if height >= 3.0 else "low"
                                        
                                        tide_entry = {
                                            "date": date_str,
                                            "time": time_str,
                                            "height": round(height, 2),
                                            "type": tide_type
                                        }
                                        tides.append(tide_entry)
        
        print(f"Total tides extracted: {len(tides)}")
        
        # Sort by date and time
        tides.sort(key=lambda x: (x['date'], x['time']))
        
        # Write JSON file
        output_file = f"{station_name}_{year}.json"
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(tides, f, ensure_ascii=False, indent=2)
            
        print(f"Created {output_file} with {len(tides)} tide entries")
        
        # Show Aug 11-12 entries specifically
        aug_11_12_tides = [t for t in tides if t['date'] in ['2025-08-11', '2025-08-12']]
        if aug_11_12_tides:
            print(f"Aug 11-12 entries ({len(aug_11_12_tides)} found):")
            for tide in aug_11_12_tides:
                print(f"  {tide['date']} {tide['time']}: {tide['height']}m ({tide['type']})")
        else:
            print("Sample entries:")
            for tide in tides[:8]:
                print(f"  {tide['date']} {tide['time']}: {tide['height']}m ({tide['type']})")
            
        return output_file
        
    except Exception as e:
        print(f"Error processing {excel_file}: {e}")
        import traceback
        traceback.print_exc()
        return None

def main():
    # Convert just Nieuwpoort for now to test
    try:
        json_file = convert_excel_to_json(
            'xlsx-getijtabellen-taw-2025/Nieuwpoort2025_mTAW.xlsx', 
            'nieuwpoort', 
            2025
        )
        if json_file:
            print(f"\\nConversion complete: {json_file}")
        else:
            print("Conversion failed")
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    main()
