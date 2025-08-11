#!/usr/bin/env python3
"""
Convert Belgian tide Excel files to JSON format for the iOS app.
Reads Excel files like Nieuwpoort2025_mTAW.xlsx and creates station_year.json files.
"""

import json
import sys
from datetime import datetime
import re

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl not installed. Run: pip3 install openpyxl")
    sys.exit(1)

def convert_excel_to_json(excel_file, station_name, year):
    """Convert an Excel file to JSON format."""
    print(f"Converting {excel_file} for {station_name} {year}...")
    
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        tides = []
        current_month = None
        
        print(f"Sheet has {ws.max_row} rows, {ws.max_column} columns")
        
        # Scan all rows looking for tide data
        for row in range(1, ws.max_row + 1):
            # Get all values in this row
            row_data = []
            for col in range(1, min(15, ws.max_column + 1)):  # Check first 15 columns
                val = ws.cell(row=row, column=col).value
                row_data.append(val)
            
            # Skip empty rows
            if all(v is None or str(v).strip() == '' for v in row_data):
                continue
                
            # Look for month headers (like "juli 2025", "august", etc.)
            month_text = str(row_data[0] or '').lower().strip()
            if any(month in month_text for month in ['januari', 'februari', 'maart', 'april', 'mei', 'juni', 
                                                    'juli', 'augustus', 'september', 'oktober', 'november', 'december',
                                                    'january', 'february', 'march', 'april', 'may', 'june',
                                                    'july', 'august', 'september', 'october', 'november', 'december']):
                # Extract month number
                month_map = {
                    'januari': 1, 'january': 1, 'februari': 2, 'february': 2, 'maart': 3, 'march': 3,
                    'april': 4, 'mei': 5, 'may': 5, 'juni': 6, 'june': 6,
                    'juli': 7, 'july': 7, 'augustus': 8, 'august': 8,
                    'september': 9, 'oktober': 10, 'october': 10, 'november': 11, 'december': 12
                }
                for month_name, month_num in month_map.items():
                    if month_name in month_text:
                        current_month = month_num
                        print(f"Found month header: {month_text} -> month {current_month}")
                        break
                continue
            
            # Look for day entries
            day_val = row_data[0]
            if isinstance(day_val, (int, float)) and 1 <= day_val <= 31 and current_month:
                day = int(day_val)
                
                # Look for time and height data in subsequent columns
                times_heights = []
                for col_idx in range(1, len(row_data)):
                    val = row_data[col_idx]
                    if val is None:
                        continue
                        
                    val_str = str(val).strip()
                    
                    # Check if it's a time (HH:MM format)
                    time_match = re.match(r'(\d{1,2}):(\d{2})', val_str)
                    if time_match:
                        hour, minute = int(time_match.group(1)), int(time_match.group(2))
                        times_heights.append(('time', f"{hour:02d}:{minute:02d}"))
                        continue
                    
                    # Check if it's a height (number with optional comma/decimal)
                    height_match = re.match(r'(\d+)[,.](\d+)', val_str)
                    if height_match:
                        height = float(f"{height_match.group(1)}.{height_match.group(2)}")
                        times_heights.append(('height', height))
                        continue
                        
                    # Check if it's just a number
                    try:
                        num_val = float(val_str.replace(',', '.'))
                        if 0 <= num_val <= 10:  # Reasonable height range
                            times_heights.append(('height', num_val))
                        elif 0 <= num_val <= 2400:  # Could be time in HHMM format
                            if num_val >= 100:
                                hour = int(num_val // 100)
                                minute = int(num_val % 100)
                                if 0 <= hour <= 23 and 0 <= minute <= 59:
                                    times_heights.append(('time', f"{hour:02d}:{minute:02d}"))
                    except ValueError:
                        pass
                
                # Pair up times and heights
                times = [item[1] for item in times_heights if item[0] == 'time']
                heights = [item[1] for item in times_heights if item[0] == 'height']
                
                # Create tide entries
                for i in range(min(len(times), len(heights))):
                    try:
                        date_str = f"{year:04d}-{current_month:02d}-{day:02d}"
                        time_str = times[i]
                        height = heights[i]
                        
                        # Determine tide type (high/low) based on height
                        tide_type = "high" if height >= 3.0 else "low"
                        
                        tide_entry = {
                            "date": date_str,
                            "time": time_str,
                            "height": round(height, 2),
                            "type": tide_type
                        }
                        tides.append(tide_entry)
                        
                    except Exception as e:
                        print(f"Error processing day {day}, time {times[i] if i < len(times) else 'N/A'}: {e}")
                        continue
                        
                if times or heights:
                    print(f"Day {day}: Found {len(times)} times, {len(heights)} heights -> {min(len(times), len(heights))} tides")
        
        print(f"Total tides extracted: {len(tides)}")
        
        # Sort by date and time
        tides.sort(key=lambda x: (x['date'], x['time']))
        
        # Write JSON file
        output_file = f"{station_name}_{year}.json"
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(tides, f, ensure_ascii=False, indent=2)
            
        print(f"Created {output_file} with {len(tides)} tide entries")
        
        # Show sample entries
        if tides:
            print("Sample entries:")
            for tide in tides[:5]:
                print(f"  {tide['date']} {tide['time']}: {tide['height']}m ({tide['type']})")
            
        return output_file
        
    except Exception as e:
        print(f"Error processing {excel_file}: {e}")
        return None

def main():
    # Station mappings - look for files in various locations
    import os
    stations = {
        'Nieuwpoort2025_mTAW.xlsx': ('nieuwpoort', 2025),
        'Antwerpen2025_mTAW.xlsx': ('antwerpen', 2025),
        'Blankenberge2025_mTAW.xlsx': ('blankenberge', 2025),
        'Oostende2025_mTAW.xlsx': ('oostende', 2025),
        'Zeebrugge2025_mTAW.xlsx': ('zeebrugge', 2025),
    }
    
    # Look for files in current dir and xlsx-getijtabellen-taw-2025 subdirectory
    search_paths = ['.', 'xlsx-getijtabellen-taw-2025']
    
    found_stations = {}
    for filename in stations.keys():
        for path in search_paths:
            full_path = os.path.join(path, filename)
            if os.path.exists(full_path):
                found_stations[full_path] = stations[filename]
                print(f"Found: {full_path}")
                break
        else:
            print(f"Warning: {filename} not found in {search_paths}")
    
    stations = found_stations
    
    converted_files = []
    
    for excel_file, (station_name, year) in stations.items():
        try:
            json_file = convert_excel_to_json(excel_file, station_name, year)
            if json_file:
                converted_files.append(json_file)
        except Exception as e:
            print(f"Failed to convert {excel_file}: {e}")
            continue
    
    print(f"\nConversion complete. Created {len(converted_files)} JSON files:")
    for file in converted_files:
        print(f"  - {file}")
    
    if converted_files:
        print(f"\nTo use in the app, copy these files to:")
        print(f"  - App bundle (add to Xcode project), OR")
        print(f"  - App Documents directory for side-loading")

if __name__ == "__main__":
    main()
