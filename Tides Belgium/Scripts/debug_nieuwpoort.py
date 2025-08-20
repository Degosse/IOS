#!/usr/bin/env python3

import openpyxl
import json
from datetime import datetime, time

def parse_time(time_val):
    """Parse time from various formats"""
    if isinstance(time_val, time):
        return time_val.strftime('%H:%M')
    elif isinstance(time_val, str):
        try:
            if ':' in time_val:
                parts = time_val.split(':')
                hours = int(parts[0])
                minutes = int(parts[1])
                return f'{hours:02d}:{minutes:02d}'
        except:
            pass
    return None

def parse_height(height_val):
    """Parse height from various formats"""
    if isinstance(height_val, (int, float)):
        return float(height_val)
    elif isinstance(height_val, str) and height_val.strip() != '-':
        try:
            return float(height_val.replace(',', '.'))
        except:
            pass
    return None

def extract_nieuwpoort_august_debug():
    excel_path = '../SourceData/xlsx-getijtabellen-taw-2025/Nieuwpoort2025_mTAW.xlsx'
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb['jul-aug']
    
    august_tides = []
    year = 2025
    month = 8  # August
    
    print(f"🔍 Extracting August {year} data from jul-aug sheet")
    print(f"Sheet dimensions: {sheet.max_row} rows x {sheet.max_column} columns")
    
    for row in range(1, sheet.max_row + 1):
        # Look for August days (second month in jul-aug sheet)
        for col in [22]:  # August days are in column 22
            day_val = sheet.cell(row=row, column=col).value
            if isinstance(day_val, (int, float)):
                day = int(day_val)
                if 20 <= day <= 21:  # Focus on days 20-21
                    print(f"\n📅 Processing August {day} (row {row})")
                    
                    try:
                        # Validate date
                        datetime(year, month, day)
                        date_str = f'{year}-{month:02d}-{day:02d}'
                        
                        # August column mapping (second month)
                        column_sets = [(24, 25, 22), (26, 27, 22)]
                        
                        # Extract from main row
                        print(f"  Main row {row}:")
                        for time_col, height_col, day_col in column_sets:
                            time_val = sheet.cell(row=row, column=time_col).value
                            height_val = sheet.cell(row=row, column=height_col).value
                            
                            time_str = parse_time(time_val)
                            height = parse_height(height_val)
                            
                            print(f"    Cols {time_col}-{height_col}: {time_val} -> {time_str}, {height_val} -> {height}")
                            
                            if time_str and height is not None:
                                tide_type = 'high' if height >= 2.5 else 'low'
                                tide = {
                                    'date': date_str,
                                    'time': time_str,
                                    'height': round(height, 2),
                                    'type': tide_type
                                }
                                august_tides.append(tide)
                                print(f"      ✅ Added: {tide}")
                        
                        # Check continuation row
                        next_row = row + 1
                        if next_row <= sheet.max_row:
                            next_day_val = sheet.cell(row=next_row, column=1).value
                            print(f"  Next row {next_row}, col1 = {next_day_val}")
                            
                            if not isinstance(next_day_val, (int, float)):
                                print(f"  Processing continuation row {next_row}:")
                                for time_col, height_col, day_col in column_sets:
                                    time_val = sheet.cell(row=next_row, column=time_col).value
                                    height_val = sheet.cell(row=next_row, column=height_col).value
                                    
                                    time_str = parse_time(time_val)
                                    height = parse_height(height_val)
                                    
                                    print(f"    Cols {time_col}-{height_col}: {time_val} -> {time_str}, {height_val} -> {height}")
                                    
                                    if time_str and height is not None:
                                        tide_type = 'high' if height >= 2.5 else 'low'
                                        tide = {
                                            'date': date_str,
                                            'time': time_str,
                                            'height': round(height, 2),
                                            'type': tide_type
                                        }
                                        august_tides.append(tide)
                                        print(f"      ✅ Added continuation: {tide}")
                            else:
                                print(f"  Skipping continuation (next row has day {next_day_val})")
                    
                    except ValueError:
                        print(f"  Invalid date: {year}-{month}-{day}")
                        continue
    
    # Sort the tides
    august_tides.sort(key=lambda x: (x['date'], x['time']))
    
    print(f"\n📋 Final August tides for days 20-21:")
    for tide in august_tides:
        print(f"  {tide['date']} {tide['time']} {tide['height']}m ({tide['type']})")
    
    return august_tides

# Run the debug extraction
tides = extract_nieuwpoort_august_debug()
