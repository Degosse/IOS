#!/usr/bin/env python3

import openpyxl
import json
from datetime import datetime, time

def parse_time(time_val):
    """Parse time from various formats"""
    if isinstance(time_val, time):
        return time_val.strftime('%H:%M')
    elif isinstance(time_val, str):
        try:
            if ':' in time_val:
                parts = time_val.split(':')
                hours = int(parts[0])
                minutes = int(parts[1])
                return f'{hours:02d}:{minutes:02d}'
        except:
            pass
    return None

def parse_height(height_val):
    """Parse height from various formats"""
    if isinstance(height_val, (int, float)):
        return float(height_val)
    elif isinstance(height_val, str) and height_val.strip() != '-':
        try:
            return float(height_val.replace(',', '.'))
        except:
            pass
    return None

def extract_station_data(station_name, year):
    """Extract data for one station"""
    
    # File name mapping
    name_map = {
        'nieuwpoort': 'Nieuwpoort',
        'blankenberge': 'Blankenberge',
        'oostende': 'Oostende',
        'zeebrugge': 'Zeebrugge'
    }
    
    excel_filename = f"{name_map[station_name]}{year}_mTAW.xlsx"
    excel_path = f'../SourceData/xlsx-getijtabellen-taw-{year}/{excel_filename}'
    
    wb = openpyxl.load_workbook(excel_path)
    all_tides = []
    
    # Process each sheet
    sheets = ['jan-feb', 'mrt-apr', 'mei-jun', 'jul-aug', 'sept-okt', 'nov-dec']
    
    for sheet_idx, sheet_name in enumerate(sheets):
        if sheet_name not in wb.sheetnames:
            continue
            
        ws = wb[sheet_name]
        
        # Process both months in the sheet
        for month_idx in range(2):
            month = (sheet_idx * 2) + month_idx + 1
            
            if month > 12:
                continue
                
            print(f"    Processing {sheet_name}, month {month}")
            
            # Find days for this month
            for row in range(1, ws.max_row + 1):
                # Look for day numbers in appropriate columns
                day_columns = [1, 8, 22] if month_idx == 0 else [8, 22]
                
                for day_col in day_columns:
                    day_val = ws.cell(row=row, column=day_col).value
                    
                    if isinstance(day_val, (int, float)):
                        day = int(day_val)
                        
                        try:
                            # Validate date
                            datetime(year, month, day)
                            date_str = f'{year}-{month:02d}-{day:02d}'
                            
                            # Column mapping based on month within sheet
                            if month_idx == 0:  # First month
                                if day_col == 1:
                                    column_sets = [(3, 4, 1), (5, 6, 1)]
                                elif day_col == 8:
                                    column_sets = [(10, 11, 8), (12, 13, 8)]
                            else:  # Second month
                                if day_col == 8:
                                    column_sets = [(17, 18, 8), (19, 20, 8)]
                                elif day_col == 22:
                                    column_sets = [(24, 25, 22), (26, 27, 22)]
                            
                            # Extract from main row
                            for time_col, height_col, _ in column_sets:
                                time_val = ws.cell(row=row, column=time_col).value
                                height_val = ws.cell(row=row, column=height_col).value
                                
                                time_str = parse_time(time_val)
                                height = parse_height(height_val)
                                
                                if time_str and height is not None:
                                    tide_type = 'high' if height >= 2.5 else 'low'
                                    all_tides.append({
                                        'date': date_str,
                                        'time': time_str,
                                        'height': round(height, 2),
                                        'type': tide_type
                                    })
                            
                            # Check continuation row
                            next_row = row + 1
                            if next_row <= ws.max_row:
                                next_day_val = ws.cell(row=next_row, column=1).value
                                if not isinstance(next_day_val, (int, float)):
                                    # Process continuation row
                                    for time_col, height_col, _ in column_sets:
                                        time_val = ws.cell(row=next_row, column=time_col).value
                                        height_val = ws.cell(row=next_row, column=height_col).value
                                        
                                        time_str = parse_time(time_val)
                                        height = parse_height(height_val)
                                        
                                        if time_str and height is not None:
                                            tide_type = 'high' if height >= 2.5 else 'low'
                                            all_tides.append({
                                                'date': date_str,
                                                'time': time_str,
                                                'height': round(height, 2),
                                                'type': tide_type
                                            })
                        
                        except ValueError:
                            # Invalid date
                            continue
    
    # Sort by date and time
    all_tides.sort(key=lambda x: (x['date'], x['time']))
    
    # Remove exact duplicates only (not the aggressive deduplication)
    unique_tides = []
    for tide in all_tides:
        if not unique_tides or unique_tides[-1] != tide:
            unique_tides.append(tide)
    
    return unique_tides

# Extract just Nieuwpoort for testing
print("🔍 EXTRACTING NIEUWPOORT 2025 WITH MINIMAL POST-PROCESSING")
print("=" * 60)

tides = extract_station_data('nieuwpoort', 2025)

# Filter for August 20-21 to check
august_tides = [t for t in tides if t['date'] in ['2025-08-20', '2025-08-21']]

print("📅 August 20-21 results:")
for tide in august_tides:
    print(f"  {tide['date']} {tide['time']} {tide['height']}m ({tide['type']})")

print(f"\nTotal tides extracted: {len(tides)}")

# Save to test file
with open('../Data/nieuwpoort_2025_fixed.json', 'w') as f:
    json.dump(tides, f, indent=2)

print("💾 Saved to nieuwpoort_2025_fixed.json")
