#!/usr/bin/env python3

import openpyxl
from datetime import datetime, time

def parse_time(time_val):
    """Parse time from various formats"""
    if isinstance(time_val, time):
        return time_val.strftime('%H:%M')
    elif isinstance(time_val, str):
        try:
            if ':' in time_val:
                parts = time_val.split(':')
                hours = int(parts[0])
                minutes = int(parts[1])
                return f'{hours:02d}:{minutes:02d}'
        except:
            pass
    return None

def parse_height(height_val):
    """Parse height from various formats"""
    if isinstance(height_val, (int, float)):
        return float(height_val)
    elif isinstance(height_val, str) and height_val.strip() != '-':
        try:
            return float(height_val.replace(',', '.'))
        except:
            pass
    return None

# Debug extraction for Nieuwpoort August 20-21
excel_path = '../SourceData/xlsx-getijtabellen-taw-2025/Nieuwpoort2025_mTAW.xlsx'
wb = openpyxl.load_workbook(excel_path)
sheet = wb['jul-aug']

print("🔍 DEBUG: Extracting August 20-21 data for Nieuwpoort")
print("=" * 60)

# Focus on rows 16-22 (around August 20-21)
for row in range(16, 23):
    print(f"\n📍 Row {row}:")
    
    # Check for day numbers
    day_col1 = sheet.cell(row=row, column=1).value
    day_col8 = sheet.cell(row=row, column=8).value  
    day_col22 = sheet.cell(row=row, column=22).value
    
    print(f"  Day values: col1={day_col1}, col8={day_col8}, col22={day_col22}")
    
    # If we found day 20 or 21
    target_day = None
    if day_col22 and str(day_col22).isdigit() and int(day_col22) in [20, 21]:
        target_day = int(day_col22)
        print(f"  ✅ Found day {target_day} in col 22")
        
        # Extract using August column mapping
        # Second month (August): cols 24-27 for days 16-31
        column_sets = [(24, 25, 22), (26, 27, 22)]
        
        for time_col, height_col, day_col in column_sets:
            time_val = sheet.cell(row=row, column=time_col).value
            height_val = sheet.cell(row=row, column=height_col).value
            
            time_str = parse_time(time_val)
            height = parse_height(height_val)
            
            print(f"    Cols {time_col}-{height_col}: time={time_val} -> {time_str}, height={height_val} -> {height}")
            
            if time_str and height is not None:
                tide_type = 'high' if height >= 2.5 else 'low'
                print(f"      ✅ TIDE: 2025-08-{target_day:02d} {time_str} {height}m ({tide_type})")
        
        # Check continuation row
        next_row = row + 1
        if next_row <= sheet.max_row:
            next_day_val = sheet.cell(row=next_row, column=1).value
            print(f"  Next row ({next_row}) col1 value: {next_day_val}")
            
            if not isinstance(next_day_val, (int, float)):
                print(f"  ✅ Processing continuation row {next_row}")
                
                for time_col, height_col, day_col in column_sets:
                    time_val = sheet.cell(row=next_row, column=time_col).value
                    height_val = sheet.cell(row=next_row, column=height_col).value
                    
                    time_str = parse_time(time_val)
                    height = parse_height(height_val)
                    
                    print(f"    Continuation cols {time_col}-{height_col}: time={time_val} -> {time_str}, height={height_val} -> {height}")
                    
                    if time_str and height is not None:
                        tide_type = 'high' if height >= 2.5 else 'low'
                        print(f"      ✅ CONTINUATION TIDE: 2025-08-{target_day:02d} {time_str} {height}m ({tide_type})")
