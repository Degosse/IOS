#!/usr/bin/env python3

import openpyxl
import json
import sys
from datetime import datetime, time
from collections import defaultdict

def parse_time(time_val):
    """Parse time from various formats"""
    if isinstance(time_val, time):
        return time_val.strftime('%H:%M')
    elif isinstance(time_val, str):
        try:
            if ':' in time_val:
                parts = time_val.split(':')
                hours = int(parts[0])
                minutes = int(parts[1])
                return f'{hours:02d}:{minutes:02d}'
        except:
            pass
    return None

def parse_height(height_val):
    """Parse height from various formats"""
    if isinstance(height_val, (int, float)):
        return float(height_val)
    elif isinstance(height_val, str) and height_val.strip() != '-':
        try:
            return float(height_val.replace(',', '.'))
        except:
            pass
    return None

def extract_station_data(station_name, year):
    """Extract data for one station with fixed column mapping"""
    
    # File name mapping
    name_map = {
        'nieuwpoort': 'Nieuwpoort',
        'blankenberge': 'Blankenberge', 
        'oostende': 'Oostende',
        'zeebrugge': 'Zeebrugge'
    }
    
    # File name mapping - handle different naming patterns per year
    if year == 2025:
        excel_filename = f"{name_map[station_name]}{year}_mTAW.xlsx"
    else:
        excel_filename = f"{name_map[station_name]}_{year}_mTAW.xlsx"
    excel_path = f'../SourceData/xlsx-getijtabellen-taw-{year}/{excel_filename}'
    
    try:
        wb = openpyxl.load_workbook(excel_path)
    except FileNotFoundError:
        print(f"  ❌ Excel file not found: {excel_filename}")
        return []
    
    all_tides = []
    
    # Process each sheet
    sheets = ['jan-feb', 'mrt-apr', 'mei-jun', 'jul-aug', 'sept-okt', 'nov-dec']
    
    for sheet_idx, sheet_name in enumerate(sheets):
        if sheet_name not in wb.sheetnames:
            continue
            
        ws = wb[sheet_name]
        months = [(sheet_idx * 2) + 1, (sheet_idx * 2) + 2]
        
        # Process both months in this sheet
        for month_idx, month in enumerate(months):
            if month > 12:
                continue
                
            print(f"    📅 Processing {sheet_name}, month {month}")
            
            # Define day columns and corresponding column sets for this month
            if month_idx == 0:  # First month (e.g., July in jul-aug)
                day_column_mappings = [
                    (1, [(3, 4), (5, 6)], range(1, 16)),      # Days 1-15: day in col 1, data in cols 3-6
                    (8, [(10, 11), (12, 13)], range(16, 32))  # Days 16-31: day in col 8, data in cols 10-13  
                ]
            else:  # Second month (e.g., August in jul-aug)
                day_column_mappings = [
                    (8, [(17, 18), (19, 20)], range(1, 16)),    # Days 1-15: day in col 8, data in cols 17-20
                    (22, [(24, 25), (26, 27)], range(16, 32))   # Days 16-31: day in col 22, data in cols 24-27
                ]
            
            # Process each day column mapping
            for day_col, time_height_pairs, valid_day_range in day_column_mappings:
                for row in range(4, ws.max_row + 1):
                    day_val = ws.cell(row=row, column=day_col).value
                    
                    if not isinstance(day_val, (int, float)) or day_val <= 0 or day_val > 31:
                        continue
                        
                    day = int(day_val)
                    
                    # Only process days that are in the valid range for this column section
                    if day not in valid_day_range:
                        continue
                    
                    try:
                        # Validate date exists
                        datetime(year, month, day)
                        date_str = f'{year}-{month:02d}-{day:02d}'
                        
                        # Extract tides from main row
                        for time_col, height_col in time_height_pairs:
                            time_val = ws.cell(row=row, column=time_col).value
                            height_val = ws.cell(row=row, column=height_col).value

                            time_str = parse_time(time_val)
                            height = parse_height(height_val)

                            if time_str and height is not None:
                                tide_type = 'high' if height >= 2.5 else 'low'
                                all_tides.append({
                                    'date': date_str,
                                    'time': time_str,
                                    'height': round(height, 2),
                                    'type': tide_type
                                })
                        
                        # Check continuation row
                        next_row = row + 1
                        if next_row <= ws.max_row:
                            next_day_val = ws.cell(row=next_row, column=1).value
                            if not isinstance(next_day_val, (int, float)):
                                # Process continuation row
                                for time_col, height_col in time_height_pairs:
                                    time_val = ws.cell(row=next_row, column=time_col).value
                                    height_val = ws.cell(row=next_row, column=height_col).value

                                    time_str = parse_time(time_val)
                                    height = parse_height(height_val)

                                    if time_str and height is not None:
                                        tide_type = 'high' if height >= 2.5 else 'low'
                                        all_tides.append({
                                            'date': date_str,
                                            'time': time_str,
                                            'height': round(height, 2),
                                            'type': tide_type
                                        })
                        
                    except ValueError:
                        # Invalid date (e.g., Feb 30), continue
                        continue
    
    # Sort and remove exact duplicates only
    all_tides.sort(key=lambda x: (x['date'], x['time']))
    
    unique_tides = []
    for tide in all_tides:
        if not unique_tides or unique_tides[-1] != tide:
            unique_tides.append(tide)
    
    print(f"  ✅ Extracted {len(unique_tides)} unique tides")
    return unique_tides

def main():
    # Get year from command line argument
    if len(sys.argv) != 2:
        print("Usage: python3 extract_year_data_fixed.py [YEAR]")
        print("Example: python3 extract_year_data_fixed.py 2025")
        sys.exit(1)
    
    try:
        year = int(sys.argv[1])
    except ValueError:
        print("Error: Year must be a number")
        sys.exit(1)
    
    # Station configurations
    stations = [
        {'name': 'blankenberge'},
        {'name': 'nieuwpoort'},
        {'name': 'oostende'}, 
        {'name': 'zeebrugge'}
    ]
    
    print(f"🗓️  EXTRACTING TIDE DATA FOR {year}")
    print("=" * 50)
    print(f"📂 Looking for Excel files in: ../SourceData/xlsx-getijtabellen-taw-{year}/")
    print(f"💾 Output will be saved to: ../Data/{year}/")
    print()
    
    # Ensure output directory exists
    import os
    output_dir = f"../Data/{year}/"
    os.makedirs(output_dir, exist_ok=True)
    
    success_count = 0
    
    for station in stations:
        station_name = station['name']
        print(f"Processing {station_name.upper()} for year {year}...")
        
        try:
            tides = extract_station_data(station_name, year)
            
            if tides:
                # Save to JSON
                output_file = f"{output_dir}{station_name}_{year}.json"
                with open(output_file, 'w') as f:
                    json.dump(tides, f, indent=2)
                
                print(f"  💾 Saved: {output_file}")
                print(f"  📅 Sample data:")
                if len(tides) >= 2:
                    print(f"    {tides[0]['date']} {tides[0]['time']}: {tides[0]['height']}m ({tides[0]['type']})")
                    print(f"    {tides[1]['date']} {tides[1]['time']}: {tides[1]['height']}m ({tides[1]['type']})")
                print()
                success_count += 1
            else:
                print(f"  ❌ No data extracted for {station_name}")
                print()
        
        except Exception as e:
            print(f"  ❌ Error processing {station_name}: {e}")
            print()
    
    print("🎉 EXTRACTION COMPLETE!")
    print(f"✅ Successfully processed {success_count}/{len(stations)} stations")
    print(f"📱 Deploy the JSON files from ../Data/{year}/ to your iOS app")
    
    if success_count < len(stations):
        print()
        print("⚠️  Some extractions failed. Check that:")
        print("   1. Excel files exist in the correct folder")
        print("   2. Excel files have the expected naming pattern")
        print("   3. Excel files have the standard Belgian tide data structure")

if __name__ == "__main__":
    main()
