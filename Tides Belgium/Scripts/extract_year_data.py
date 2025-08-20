#!/usr/bin/env python3
"""
Belgian Tide Data Extraction Tool
Master script to extract data from Excel files for any year
Usage: python3 extract_year_data.py [YEAR]
Example: python3 extract_year_data.py 2027
"""

import sys
import openpyxl
import json
import os
from datetime import datetime

def parse_time(time_val):
    """Parse time from Excel cell value"""
    if time_val is None:
        return None
    if hasattr(time_val, 'hour'):
        return f'{time_val.hour:02d}:{time_val.minute:02d}'
    time_str = str(time_val).strip()
    if ':' in time_str:
        parts = time_str.split(':')
        if len(parts) >= 2:
            try:
                hour = int(parts[0])
                minute = int(parts[1])
                return f'{hour:02d}:{minute:02d}'
            except ValueError:
                pass
    return None

def parse_height(height_val):
    """Parse height from Excel cell value"""
    if height_val is None:
        return None
    try:
        if isinstance(height_val, (int, float)):
            return float(height_val)
        height_str = str(height_val).replace(',', '.')
        return float(height_str)
    except ValueError:
        return None

def get_months_from_sheet_name(sheet_name):
    """Convert Dutch sheet names to month numbers"""
    month_map = {
        'jan-feb': [1, 2],
        'mrt-apr': [3, 4], 
        'mei-jun': [5, 6],
        'jul-aug': [7, 8],
        'sept-okt': [9, 10],
        'nov-dec': [11, 12]
    }
    return month_map.get(sheet_name.lower(), [])

def extract_station_data(excel_path, station_name, year):
    """Extract full year tide data from a station's Excel file"""
    print(f"Processing {station_name.upper()} for year {year}...")
    
    if not os.path.exists(excel_path):
        print(f"  ❌ Excel file not found: {excel_path}")
        return []
    
    try:
        wb = openpyxl.load_workbook(excel_path)
    except Exception as e:
        print(f"  ❌ Error loading Excel file: {e}")
        return []
    
    all_tides = []
    
    for sheet_name in wb.sheetnames:
        if '-' not in sheet_name.lower():
            continue
            
        months = get_months_from_sheet_name(sheet_name)
        if not months:
            continue
            
        print(f"  📋 Processing sheet: {sheet_name}")
        ws = wb[sheet_name]
        
        # Process both months in this sheet  
        for month_idx, month in enumerate(months):
            print(f"    📅 Processing month {month}")
            
            # Determine day columns and column sets for this month
            if month_idx == 0:  # First month (e.g., July in jul-aug)
                # Days 1-15 in col 1, days 16-31 in col 8
                day_columns_map = {1: [(3, 4, 1), (5, 6, 1)], 8: [(10, 11, 8), (12, 13, 8)]}
            else:  # Second month (e.g., August in jul-aug)
                # Days 1-15 in col 8, days 16-31 in col 22  
                day_columns_map = {8: [(17, 18, 8), (19, 20, 8)], 22: [(24, 25, 22), (26, 27, 22)]}
            
            # Process each day column for this month
            for day_col, column_sets in day_columns_map.items():
                for row in range(4, ws.max_row + 1):
                    day_val = ws.cell(row=row, column=day_col).value
                    
                    if not isinstance(day_val, (int, float)) or day_val <= 0 or day_val > 31:
                        continue
                        
                    day = int(day_val)
                    try:
                        # Validate date exists
                        datetime(year, month, day)
                        date_str = f'{year}-{month:02d}-{day:02d}'
                        
                        # Extract tides from main row
                        for time_col, height_col, day_col in column_sets:
                        # Day may be repeated in alternative columns depending on block
                        day_source = ws.cell(row=row, column=day_col).value
                        if isinstance(day_source, (int, float)) and 1 <= int(day_source) <= 31:
                            day_for_block = int(day_source)
                        else:
                            # Fallback to primary day detected earlier
                            day_for_block = day
                        date_str = f'{year}-{month:02d}-{day_for_block:02d}'

                        time_val = ws.cell(row=row, column=time_col).value
                        height_val = ws.cell(row=row, column=height_col).value

                        time_str = parse_time(time_val)
                        height = parse_height(height_val)

                        if time_str and height is not None:
                            tide_type = 'high' if height >= 2.5 else 'low'
                            all_tides.append({
                                'date': date_str,
                                'time': time_str,
                                'height': round(height, 2),
                                'type': tide_type
                            })
                    
                    # Check continuation row
                    next_row = row + 1
                    if next_row <= ws.max_row:
                        next_day_val = ws.cell(row=next_row, column=1).value
                        if not isinstance(next_day_val, (int, float)):
                            for time_col, height_col, day_col in column_sets:
                                day_source = ws.cell(row=row, column=day_col).value
                                if isinstance(day_source, (int, float)) and 1 <= int(day_source) <= 31:
                                    day_for_block = int(day_source)
                                else:
                                    day_for_block = day
                                date_str = f'{year}-{month:02d}-{day_for_block:02d}'

                                time_val = ws.cell(row=next_row, column=time_col).value
                                height_val = ws.cell(row=next_row, column=height_col).value

                                time_str = parse_time(time_val)
                                height = parse_height(height_val)

                                if time_str and height is not None:
                                    tide_type = 'high' if height >= 2.5 else 'low'
                                    all_tides.append({
                                        'date': date_str,
                                        'time': time_str,
                                        'height': round(height, 2),
                                        'type': tide_type
                                    })
                    
                except ValueError:
                    # Invalid date (e.g., Feb 30), continue
                    continue
    
    # Sort and deduplicate
    all_tides.sort(key=lambda x: (x['date'], x['time']))
    
    unique_tides = []
    seen = set()
    for tide in all_tides:
        key = f"{tide['date']}_{tide['time']}_{tide['height']}"
        if key not in seen:
            seen.add(key)
            unique_tides.append(tide)
    
    # Post-process: collapse near-duplicates within the same day and cap to 4 tides/day
    final_tides = []
    from collections import defaultdict
    by_date: dict = defaultdict(list)
    for t in unique_tides:
        by_date[t['date']].append(t)

    def time_to_minutes(tstr: str) -> int:
        h, m = tstr.split(':')
        return int(h) * 60 + int(m)

    def dedupe_cluster(entries, is_high: bool):
        if not entries:
            return []
        entries = sorted(entries, key=lambda x: time_to_minutes(x['time']))
        result = []
        used = [False] * len(entries)
        T_WINDOW = 60  # minutes
        H_WINDOW = 0.4   # meters
        for i, a in enumerate(entries):
            if used[i]:
                continue
            cluster = [i]
            for j in range(i + 1, len(entries)):
                if used[j]:
                    continue
                b = entries[j]
                if abs(time_to_minutes(a['time']) - time_to_minutes(b['time'])) <= T_WINDOW \
                   and abs(a['height'] - b['height']) <= H_WINDOW:
                    cluster.append(j)
            # pick representative: extreme height for highs, opposite for lows
            pick = entries[cluster[0]]
            for k in cluster[1:]:
                cand = entries[k]
                if is_high:
                    if cand['height'] > pick['height']:
                        pick = cand
                else:
                    if cand['height'] < pick['height']:
                        pick = cand
                used[k] = True
            used[i] = True
            result.append(pick)
        return result

    for date, tides in by_date.items():
        # Separate by type using the precomputed 'type'
        highs = [x for x in tides if x['type'] == 'high']
        lows = [x for x in tides if x['type'] == 'low']

        highs = dedupe_cluster(highs, True)
        lows = dedupe_cluster(lows, False)

        # Limit to at most two highs and two lows per day (keep earliest two by time)
        highs = sorted(highs, key=lambda x: time_to_minutes(x['time']))[:2]
        lows = sorted(lows, key=lambda x: time_to_minutes(x['time']))[:2]

        merged = sorted(highs + lows, key=lambda x: time_to_minutes(x['time']))

        # Enforce alternation: remove consecutive same-type entries, keep more extreme/earlier
        alt = []
        for e in merged:
            if not alt:
                alt.append(e)
                continue
            if alt[-1]['type'] == e['type']:
                # Same type back-to-back, keep more extreme (higher high or lower low)
                if e['type'] == 'high':
                    alt[-1] = e if e['height'] > alt[-1]['height'] else alt[-1]
                else:
                    alt[-1] = e if e['height'] < alt[-1]['height'] else alt[-1]
            else:
                alt.append(e)

        final_tides.extend(alt[:4])  # cap at 4 per day just in case

    final_tides.sort(key=lambda x: (x['date'], x['time']))

    print(f"  ✅ Extracted {len(unique_tides)} unique tides")
    print(f"  🔎 After per-day dedupe/cap: {len(final_tides)} entries")
    return final_tides

def main():
    # Get year from command line argument
    if len(sys.argv) != 2:
        print("Usage: python3 extract_year_data.py [YEAR]")
        print("Example: python3 extract_year_data.py 2027")
        sys.exit(1)
    
    try:
        year = int(sys.argv[1])
    except ValueError:
        print("Error: Year must be a number")
        sys.exit(1)
    
    # Station configurations
    stations = [
        {'name': 'blankenberge'},
        {'name': 'nieuwpoort'},
        {'name': 'oostende'},
        {'name': 'zeebrugge'}
    ]
    
    excel_folder = f"../SourceData/xlsx-getijtabellen-taw-{year}"
    output_folder = f"../Data/{year}"
    
    # Create output folder
    os.makedirs(output_folder, exist_ok=True)
    
    print(f"🗓️  EXTRACTING TIDE DATA FOR {year}")
    print("=" * 50)
    print(f"📂 Looking for Excel files in: {excel_folder}/")
    print(f"💾 Output will be saved to: {output_folder}/")
    print()
    
    successful_extractions = 0
    
    for station in stations:
        # Try different filename patterns for different years
        station_name_title = station['name'].title()
        possible_patterns = [
            f"{station_name_title}{year}_mTAW.xlsx",  # 2025 format
            f"{station_name_title}_{year}_mTAW.xlsx",  # 2026 format
        ]
        
        excel_path = None
        for pattern in possible_patterns:
            potential_path = os.path.join(excel_folder, pattern)
            if os.path.exists(potential_path):
                excel_path = potential_path
                break
        
        if not excel_path:
            print(f"Processing {station['name'].upper()} for year {year}...")
            print(f"  ❌ Excel file not found. Tried patterns:")
            for pattern in possible_patterns:
                print(f"     - {pattern}")
            print(f"  ❌ No data extracted for {station['name']}")
            print()
            continue
        
        tides = extract_station_data(excel_path, station['name'], year)
        
        if tides:
            output_file = os.path.join(output_folder, f"{station['name']}_{year}.json")
            
            with open(output_file, 'w') as f:
                json.dump(tides, f, indent=2)
            
            print(f"  💾 Saved: {output_file}")
            successful_extractions += 1
            
            # Show sample data
            sample_tides = [t for t in tides if t['date'].endswith('-01-01') or t['date'].endswith('-08-15')]
            if sample_tides:
                print(f"  📅 Sample data:")
                for tide in sample_tides[:2]:
                    print(f"    {tide['date']} {tide['time']}: {tide['height']}m ({tide['type']})")
        else:
            print(f"  ❌ No data extracted for {station['name']}")
        
        print()
    
    print("🎉 EXTRACTION COMPLETE!")
    print(f"✅ Successfully processed {successful_extractions}/5 stations")
    print(f"📱 Deploy the JSON files from {output_folder}/ to your iOS app")
    
    if successful_extractions < 5:
        print("\n⚠️  Some extractions failed. Check that:")
        print("   1. Excel files exist in the correct folder")
        print("   2. Excel files have the expected naming pattern")
        print("   3. Excel files have the standard Belgian tide data structure")

if __name__ == "__main__":
    main()
