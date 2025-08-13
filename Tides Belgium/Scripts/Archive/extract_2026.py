#!/usr/bin/env python3
"""Extract 2026 data for future use"""

import openpyxl
import json
import os
from datetime import datetime

def parse_time(time_val):
    if time_val is None:
        return None
    if hasattr(time_val, 'hour'):
        return f'{time_val.hour:02d}:{time_val.minute:02d}'
    time_str = str(time_val).strip()
    if ':' in time_str:
        parts = time_str.split(':')
        if len(parts) >= 2:
            try:
                hour = int(parts[0])
                minute = int(parts[1])
                return f'{hour:02d}:{minute:02d}'
            except ValueError:
                pass
    return None

def parse_height(height_val):
    if height_val is None:
        return None
    try:
        if isinstance(height_val, (int, float)):
            return float(height_val)
        height_str = str(height_val).replace(',', '.')
        return float(height_str)
    except ValueError:
        return None

def get_months_from_sheet_name(sheet_name):
    month_map = {
        'jan-feb': [1, 2],
        'mrt-apr': [3, 4], 
        'mei-jun': [5, 6],
        'jul-aug': [7, 8],
        'sept-okt': [9, 10],
        'nov-dec': [11, 12]
    }
    return month_map.get(sheet_name.lower(), [])

def extract_station_full_year(excel_path, station_name, year):
    print(f"Processing {station_name.upper()} for year {year}...")
    
    if not os.path.exists(excel_path):
        print(f"Excel file not found: {excel_path}")
        return []
    
    wb = openpyxl.load_workbook(excel_path)
    all_tides = []
    
    for sheet_name in wb.sheetnames:
        if '-' not in sheet_name.lower():
            continue
            
        months = get_months_from_sheet_name(sheet_name)
        if not months:
            continue
            
        print(f"  Processing sheet: {sheet_name}")
        ws = wb[sheet_name]
        
        for row in range(4, ws.max_row + 1):
            day_val = ws.cell(row=row, column=1).value
            
            if not isinstance(day_val, (int, float)) or day_val <= 0 or day_val > 31:
                continue
                
            day = int(day_val)
            
            for month_idx, month in enumerate(months):
                try:
                    datetime(year, month, day)
                    date_str = f'{year}-{month:02d}-{day:02d}'
                    
                    if month_idx == 0:
                        tide_columns = [(3, 4), (5, 6)]
                    else:
                        tide_columns = [(10, 11)]
                    
                    for time_col, height_col in tide_columns:
                        time_val = ws.cell(row=row, column=time_col).value
                        height_val = ws.cell(row=row, column=height_col).value
                        
                        time_str = parse_time(time_val)
                        height = parse_height(height_val)
                        
                        if time_str and height is not None:
                            tide_type = 'high' if height >= 2.5 else 'low'
                            all_tides.append({
                                'date': date_str,
                                'time': time_str,
                                'height': round(height, 2),
                                'type': tide_type
                            })
                    
                    # Continuation row
                    next_row = row + 1
                    if next_row <= ws.max_row:
                        next_day_val = ws.cell(row=next_row, column=1).value
                        if not isinstance(next_day_val, (int, float)):
                            for time_col, height_col in tide_columns:
                                time_val = ws.cell(row=next_row, column=time_col).value
                                height_val = ws.cell(row=next_row, column=height_col).value
                                
                                time_str = parse_time(time_val)
                                height = parse_height(height_val)
                                
                                if time_str and height is not None:
                                    tide_type = 'high' if height >= 2.5 else 'low'
                                    all_tides.append({
                                        'date': date_str,
                                        'time': time_str,
                                        'height': round(height, 2),
                                        'type': tide_type
                                    })
                    
                except ValueError:
                    continue
    
    # Sort and deduplicate
    all_tides.sort(key=lambda x: (x['date'], x['time']))
    unique_tides = []
    seen = set()
    for tide in all_tides:
        key = f"{tide['date']}_{tide['time']}_{tide['height']}"
        if key not in seen:
            seen.add(key)
            unique_tides.append(tide)
    
    print(f"  Extracted {len(unique_tides)} unique tides")
    return unique_tides

# Process 2026 data
stations_2026 = [
    {'name': 'nieuwpoort', 'file': 'xlsx-getijtabellen-taw-2026/Nieuwpoort_2026_mTAW.xlsx'},
    {'name': 'oostende', 'file': 'xlsx-getijtabellen-taw-2026/Oostende_2026_mTAW.xlsx'},
    {'name': 'blankenberge', 'file': 'xlsx-getijtabellen-taw-2026/Blankenberge_2026_mTAW.xlsx'},
    {'name': 'zeebrugge', 'file': 'xlsx-getijtabellen-taw-2026/Zeebrugge_2026_mTAW.xlsx'},
    {'name': 'antwerpen', 'file': 'xlsx-getijtabellen-taw-2026/Antwerpen_2026_mTAW.xlsx'},
]

print("PROCESSING YEAR 2026")
print("=" * 50)

for station in stations_2026:
    tides = extract_station_full_year(station['file'], station['name'], 2026)
    
    if tides:
        output_file = f"{station['name']}_2026.json"
        with open(output_file, 'w') as f:
            json.dump(tides, f, indent=2)
        
        print(f"  Saved: {output_file}")
        print()

print("2026 EXTRACTION COMPLETE!")
print("These JSON files will be automatically used when the app detects the year has changed to 2026.")
