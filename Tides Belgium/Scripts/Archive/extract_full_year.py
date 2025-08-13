#!/usr/bin/env python3
"""Extract comprehensive tide data from Excel files for the entire year 2025"""

import openpyxl
import json
from datetime import datetime, timedelta

def parse_time(time_val):
    if time_val is None:
        return None
    if hasattr(time_val, 'hour'):
        return f'{time_val.hour:02d}:{time_val.minute:02d}'
    time_str = str(time_val).strip()
    if ':' in time_str:
        parts = time_str.split(':')
        if len(parts) >= 2:
            try:
                hour = int(parts[0])
                minute = int(parts[1])
                return f'{hour:02d}:{minute:02d}'
            except ValueError:
                pass
    return None

def parse_height(height_val):
    if height_val is None:
        return None
    try:
        if isinstance(height_val, (int, float)):
            return float(height_val)
        height_str = str(height_val).replace(',', '.')
        return float(height_str)
    except ValueError:
        return None

def get_month_from_sheet_name(sheet_name):
    """Convert sheet name to month numbers"""
    month_map = {
        'jan-feb': [1, 2],
        'mar-apr': [3, 4], 
        'may-jun': [5, 6],
        'jul-aug': [7, 8],
        'sep-oct': [9, 10],
        'nov-dec': [11, 12]
    }
    return month_map.get(sheet_name.lower(), [])

def extract_full_year_data(excel_path, station_name, year=2025):
    """Extract full year tide data from a station's Excel file"""
    print(f"Processing {station_name} for {year}...")
    
    wb = openpyxl.load_workbook(excel_path)
    tides = []
    
    # Process all sheets
    for sheet_name in wb.sheetnames:
        if '-' not in sheet_name.lower():
            continue
            
        print(f"  Processing sheet: {sheet_name}")
        ws = wb[sheet_name]
        months = get_month_from_sheet_name(sheet_name)
        
        if not months:
            continue
            
        for row in range(1, ws.max_row + 1):
            day_val = ws.cell(row=row, column=1).value
            if not isinstance(day_val, (int, float)) or day_val <= 0 or day_val > 31:
                continue
                
            day = int(day_val)
            
            # Try each month for this sheet
            for month in months:
                try:
                    date_str = f'{year}-{month:02d}-{day:02d}'
                    # Validate the date
                    datetime.strptime(date_str, '%Y-%m-%d')
                    
                    # Extract tides from main row (columns 3,4 and 5,6)
                    for time_col, height_col in [(3, 4), (5, 6)]:
                        time_val = ws.cell(row=row, column=time_col).value
                        height_val = ws.cell(row=row, column=height_col).value
                        
                        time_str = parse_time(time_val)
                        height = parse_height(height_val)
                        
                        if time_str and height is not None:
                            tide_type = 'high' if height >= 3.0 else 'low'
                            tides.append({
                                'date': date_str,
                                'time': time_str,
                                'height': round(height, 2),
                                'type': tide_type
                            })
                    
                    # Check continuation row for more tides
                    next_row = row + 1
                    if next_row <= ws.max_row:
                        next_day_val = ws.cell(row=next_row, column=1).value
                        if not isinstance(next_day_val, (int, float)):
                            for time_col, height_col in [(3, 4), (5, 6)]:
                                time_val = ws.cell(row=next_row, column=time_col).value
                                height_val = ws.cell(row=next_row, column=height_col).value
                                
                                time_str = parse_time(time_val)
                                height = parse_height(height_val)
                                
                                if time_str and height is not None:
                                    tide_type = 'high' if height >= 3.0 else 'low'
                                    tides.append({
                                        'date': date_str,
                                        'time': time_str,
                                        'height': round(height, 2),
                                        'type': tide_type
                                    })
                    
                    break  # Found valid month for this day
                    
                except ValueError:
                    # Invalid date, try next month
                    continue
    
    # Sort by date and time
    tides.sort(key=lambda x: (x['date'], x['time']))
    
    # Remove duplicates
    unique_tides = []
    seen = set()
    for tide in tides:
        key = f"{tide['date']}_{tide['time']}_{tide['height']}"
        if key not in seen:
            seen.add(key)
            unique_tides.append(tide)
    
    return unique_tides

# Extract data for all stations
stations = [
    {'name': 'nieuwpoort', 'file': 'xlsx-getijtabellen-taw-2025/Nieuwpoort2025_mTAW.xlsx'},
    {'name': 'oostende', 'file': 'xlsx-getijtabellen-taw-2025/Oostende2025_mTAW.xlsx'},
    {'name': 'blankenberge', 'file': 'xlsx-getijtabellen-taw-2025/Blankenberge2025_mTAW.xlsx'},
    {'name': 'zeebrugge', 'file': 'xlsx-getijtabellen-taw-2025/Zeebrugge2025_mTAW.xlsx'},
    {'name': 'antwerpen', 'file': 'xlsx-getijtabellen-taw-2025/Antwerpen2025_mTAW.xlsx'},
]

for station in stations:
    try:
        tides = extract_full_year_data(station['file'], station['name'])
        
        output_file = f"{station['name']}_2025_full.json"
        with open(output_file, 'w') as f:
            json.dump(tides, f, indent=2)
        
        print(f"✓ Created {output_file} with {len(tides)} tides")
        
        # Show some sample data around current date
        print(f"Sample data around Aug 11-13:")
        for tide in tides:
            if '2025-08-1' in tide['date']:
                print(f"  {tide['date']} {tide['time']}: {tide['height']}m ({tide['type']})")
        print()
        
    except Exception as e:
        print(f"✗ Error processing {station['name']}: {e}")
        print()

print("Full year extraction complete!")
