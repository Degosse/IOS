#!/usr/bin/env python3

import openpyxl

# Debug the exact column mapping for August 20 in Nieuwpoort
excel_path = '../SourceData/xlsx-getijtabellen-taw-2025/Nieuwpoort2025_mTAW.xlsx'
wb = openpyxl.load_workbook(excel_path)
sheet = wb['jul-aug']

print("🔍 DEBUGGING EXACT COLUMN MAPPING FOR AUGUST 20")
print("=" * 50)

# Find August 20 in the sheet
for row in range(1, sheet.max_row + 1):
    # Check different columns where day 20 might appear
    for col in [1, 8, 22]:
        day_val = sheet.cell(row=row, column=col).value
        if day_val == 20:
            print(f"\nFound day 20 at row {row}, column {col}")
            
            # Show context around this position
            print(f"Row {row} context:")
            for c in range(max(1, col-3), min(sheet.max_column + 1, col+10)):
                val = sheet.cell(row=row, column=c).value
                print(f"  Col {c:2}: {val}")
            
            # Determine which month this represents by looking at headers
            print(f"\nDetermining month context...")
            
            # Look for month indicators in the area
            for header_row in range(max(1, row-10), row):
                for header_col in range(max(1, col-5), min(sheet.max_column + 1, col+15)):
                    header_val = sheet.cell(row=header_row, column=header_col).value
                    if header_val and isinstance(header_val, str):
                        if 'juli' in header_val.lower() or 'july' in header_val.lower():
                            print(f"    Found July indicator at row {header_row}, col {header_col}: '{header_val}'")
                        elif 'augustus' in header_val.lower() or 'august' in header_val.lower():
                            print(f"    Found August indicator at row {header_row}, col {header_col}: '{header_val}'")
            
            # Based on column position, determine the correct extraction logic
            if col == 22:  # This should be August data
                print(f"\n✅ Column 22 suggests August data - using August column mapping")
                print(f"    Expected tide columns: 24-25 (time-height), 26-27 (time-height)")
                
                time1 = sheet.cell(row=row, column=24).value
                height1 = sheet.cell(row=row, column=25).value
                time2 = sheet.cell(row=row, column=26).value
                height2 = sheet.cell(row=row, column=27).value
                
                print(f"    Main row data: {time1} {height1}m, {time2} {height2}m")
                
                # Check continuation row
                cont_row = row + 1
                cont_time1 = sheet.cell(row=cont_row, column=24).value
                cont_height1 = sheet.cell(row=cont_row, column=25).value
                cont_time2 = sheet.cell(row=cont_row, column=26).value
                cont_height2 = sheet.cell(row=cont_row, column=27).value
                
                print(f"    Continuation row data: {cont_time1} {cont_height1}m, {cont_time2} {cont_height2}m")
                
            elif col == 8:  # This might be July data being misread as August
                print(f"\n⚠️  Column 8 might be July data misinterpreted as August")
                print(f"    July column mapping would be: 10-11, 12-13")
                
                time1 = sheet.cell(row=row, column=10).value
                height1 = sheet.cell(row=row, column=11).value
                time2 = sheet.cell(row=row, column=12).value
                height2 = sheet.cell(row=row, column=13).value
                
                print(f"    July interpretation: {time1} {height1}m, {time2} {height2}m")
