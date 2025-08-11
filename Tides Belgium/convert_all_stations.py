#!/usr/bin/env python3
"""Convert all Belgian tide stations Excel files to JSON format for Aug 11-12, 2025"""

import openpyxl
import json

def parse_time(time_val):
    if time_val is None:
        return None
    if hasattr(time_val, 'hour'):
        return f'{time_val.hour:02d}:{time_val.minute:02d}'
    time_str = str(time_val).strip()
    if ':' in time_str:
        parts = time_str.split(':')
        if len(parts) >= 2:
            try:
                hour = int(parts[0])
                minute = int(parts[1])
                return f'{hour:02d}:{minute:02d}'
            except ValueError:
                pass
    return None

def parse_height(height_val):
    if height_val is None:
        return None
    try:
        if isinstance(height_val, (int, float)):
            return float(height_val)
        height_str = str(height_val).replace(',', '.')
        return float(height_str)
    except ValueError:
        return None

def extract_station_data(excel_path, station_name):
    """Extract Aug 11-12 tide data from a station's Excel file"""
    print(f"Processing {station_name}...")
    
    wb = openpyxl.load_workbook(excel_path)
    ws = wb['jul-aug']
    
    tides = []
    
    for row in range(1, ws.max_row + 1):
        day_val = ws.cell(row=row, column=1).value
        if isinstance(day_val, (int, float)) and day_val in [11, 12]:
            day = int(day_val)
            date_str = f'2025-08-{day:02d}'
            
            print(f'  Processing August {day}...')
            
            # Extract tides from main row (columns 3,4 and 5,6)
            for time_col, height_col in [(3, 4), (5, 6)]:
                time_val = ws.cell(row=row, column=time_col).value
                height_val = ws.cell(row=row, column=height_col).value
                
                time_str = parse_time(time_val)
                height = parse_height(height_val)
                
                if time_str and height is not None:
                    tide_type = 'high' if height >= 3.0 else 'low'
                    print(f'    {time_str}: {height}m ({tide_type})')
                    tides.append({
                        'date': date_str,
                        'time': time_str,
                        'height': round(height, 2),
                        'type': tide_type
                    })
            
            # Check continuation row
            next_row = row + 1
            if next_row <= ws.max_row:
                next_day_val = ws.cell(row=next_row, column=1).value
                if not isinstance(next_day_val, (int, float)):
                    print('    Continuation row found')
                    for time_col, height_col in [(3, 4), (5, 6)]:
                        time_val = ws.cell(row=next_row, column=time_col).value
                        height_val = ws.cell(row=next_row, column=height_col).value
                        
                        time_str = parse_time(time_val)
                        height = parse_height(height_val)
                        
                        if time_str and height is not None:
                            tide_type = 'high' if height >= 3.0 else 'low'
                            print(f'      {time_str}: {height}m ({tide_type})')
                            tides.append({
                                'date': date_str,
                                'time': time_str,
                                'height': round(height, 2),
                                'type': tide_type
                            })
    
    # Sort by date and time
    tides.sort(key=lambda x: (x['date'], x['time']))
    
    return tides

# Station configurations
stations = [
    {'name': 'nieuwpoort', 'file': 'xlsx-getijtabellen-taw-2025/Nieuwpoort2025_mTAW.xlsx'},
    {'name': 'oostende', 'file': 'xlsx-getijtabellen-taw-2025/Oostende2025_mTAW.xlsx'},
    {'name': 'blankenberge', 'file': 'xlsx-getijtabellen-taw-2025/Blankenberge2025_mTAW.xlsx'},
    {'name': 'zeebrugge', 'file': 'xlsx-getijtabellen-taw-2025/Zeebrugge2025_mTAW.xlsx'},
    {'name': 'antwerpen', 'file': 'xlsx-getijtabellen-taw-2025/Antwerpen2025_mTAW.xlsx'},
]

# Process all stations
for station in stations:
    try:
        tides = extract_station_data(station['file'], station['name'])
        
        output_file = f"{station['name']}_2025.json"
        with open(output_file, 'w') as f:
            json.dump(tides, f, indent=2)
        
        print(f"✓ Created {output_file} with {len(tides)} tides")
        for tide in tides:
            print(f"  {tide['date']} {tide['time']}: {tide['height']}m ({tide['type']})")
        print()
        
    except Exception as e:
        print(f"✗ Error processing {station['name']}: {e}")
        print()

print("All stations processed!")
