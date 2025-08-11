#!/usr/bin/env python3
"""
Convert Belgian tide Excel files to JSON format for the iOS app.
Handles the specific format used in the mTAW Excel files.
"""

import json
import sys
import os
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl not installed. Run: pip3 install openpyxl")
    sys.exit(1)

def parse_time(time_val):
    """Parse time from various formats in the Excel."""
    if time_val is None:
        return None
    
    time_str = str(time_val).strip()
    
    # Handle time objects from Excel
    if hasattr(time_val, 'hour'):
        return f"{time_val.hour:02d}:{time_val.minute:02d}"
    
    # Handle HH:MM:SS format
    if ':' in time_str:
        parts = time_str.split(':')
        if len(parts) >= 2:
            try:
                hour = int(parts[0])
                minute = int(parts[1])
                return f"{hour:02d}:{minute:02d}"
            except ValueError:
                pass
    
    return None

def parse_height(height_val):
    """Parse height from Excel format."""
    if height_val is None:
        return None
    
    try:
        if isinstance(height_val, (int, float)):
            return float(height_val)
        
        height_str = str(height_val).replace(',', '.')
        return float(height_str)
    except ValueError:
        return None

def convert_excel_to_json(excel_file, station_name, year):
    """Convert an Excel file to JSON format."""
    print(f"Converting {excel_file} for {station_name} {year}...")
    
    try:
        wb = openpyxl.load_workbook(excel_file)
        
        # Try to find a sheet with data (sometimes it's not the first sheet)
        sheets_to_try = [wb.active]  # Start with active sheet
        if wb.sheetnames:
            for sheet_name in wb.sheetnames:
                if sheet_name not in [s.title for s in sheets_to_try]:
                    sheets_to_try.append(wb[sheet_name])
        
        tides = []
        
        for ws in sheets_to_try:
            print(f"Trying sheet: {ws.title}")
            
            current_month = None  # We'll need to determine this from context
            
            # The Excel format appears to have:
            # Column 1: Day number
            # Column 3: High water time, Column 4: High water height
            # Column 5: Low water time, Column 6: Low water height
            # Column 10: Another high water time, etc.
            
            for row in range(1, ws.max_row + 1):
                # Get day from column 1
                day_val = ws.cell(row=row, column=1).value
                
                if isinstance(day_val, (int, float)) and 1 <= day_val <= 31:
                    day = int(day_val)
                    
                    # For now, let's assume we're looking at November-December based on sheet name
                    # You may need to adjust this logic based on the actual sheet structure
                    if 'nov' in ws.title.lower():
                        current_month = 11
                    elif 'dec' in ws.title.lower():
                        current_month = 12
                    elif 'aug' in ws.title.lower() or 'august' in ws.title.lower():
                        current_month = 8
                    elif 'jul' in ws.title.lower() or 'july' in ws.title.lower():
                        current_month = 7
                    elif 'sep' in ws.title.lower() or 'september' in ws.title.lower():
                        current_month = 9
                    elif 'oct' in ws.title.lower() or 'october' in ws.title.lower():
                        current_month = 10
                    else:
                        # Default to August for testing
                        current_month = 8
                    
                    if current_month is None:
                        continue
                        
                    date_str = f"{year:04d}-{current_month:02d}-{day:02d}"
                    
                    # Extract tide data from this row
                    tide_columns = [
                        (3, 4, 'high'),   # Column 3=time, 4=height, high tide
                        (5, 6, 'low'),    # Column 5=time, 6=height, low tide  
                        (10, 11, 'high'), # Column 10=time, 11=height, high tide
                        (12, 13, 'low'),  # Column 12=time, 13=height, low tide
                    ]
                    
                    for time_col, height_col, tide_type in tide_columns:
                        time_val = ws.cell(row=row, column=time_col).value
                        height_val = ws.cell(row=row, column=height_col).value
                        
                        time_str = parse_time(time_val)
                        height = parse_height(height_val)
                        
                        if time_str and height is not None:
                            tide_entry = {
                                "date": date_str,
                                "time": time_str,
                                "height": round(height, 2),
                                "type": tide_type
                            }
                            tides.append(tide_entry)
                            
                    # Also check the next row for continuation of same day
                    next_row = row + 1
                    if next_row <= ws.max_row:
                        # If next row doesn't have a day number, it might be continuation
                        next_day_val = ws.cell(row=next_row, column=1).value
                        if not isinstance(next_day_val, (int, float)):
                            # This row continues the same day
                            for time_col, height_col, tide_type in tide_columns:
                                time_val = ws.cell(row=next_row, column=time_col).value
                                height_val = ws.cell(row=next_row, column=height_col).value
                                
                                time_str = parse_time(time_val)
                                height = parse_height(height_val)
                                
                                if time_str and height is not None:
                                    tide_entry = {
                                        "date": date_str,
                                        "time": time_str,
                                        "height": round(height, 2),
                                        "type": tide_type
                                    }
                                    tides.append(tide_entry)
            
            if tides:  # Found data in this sheet
                break
        
        print(f"Total tides extracted: {len(tides)}")
        
        # Sort by date and time
        tides.sort(key=lambda x: (x['date'], x['time']))
        
        # Write JSON file
        output_file = f"{station_name}_{year}.json"
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(tides, f, ensure_ascii=False, indent=2)
            
        print(f"Created {output_file} with {len(tides)} tide entries")
        
        # Show sample entries for Aug 11-12 specifically
        aug_11_12_tides = [t for t in tides if t['date'] in ['2025-08-11', '2025-08-12']]
        if aug_11_12_tides:
            print(f"Aug 11-12 entries ({len(aug_11_12_tides)} found):")
            for tide in aug_11_12_tides:
                print(f"  {tide['date']} {tide['time']}: {tide['height']}m ({tide['type']})")
        else:
            print("Sample entries:")
            for tide in tides[:8]:
                print(f"  {tide['date']} {tide['time']}: {tide['height']}m ({tide['type']})")
            
        return output_file
        
    except Exception as e:
        print(f"Error processing {excel_file}: {e}")
        import traceback
        traceback.print_exc()
        return None

def main():
    # Station mappings - look for files in various locations
    stations = {
        'Nieuwpoort2025_mTAW.xlsx': ('nieuwpoort', 2025),
        'Antwerpen2025_mTAW.xlsx': ('antwerpen', 2025),
        'Blankenberge2025_mTAW.xlsx': ('blankenberge', 2025),
        'Oostende2025_mTAW.xlsx': ('oostende', 2025),
        'Zeebrugge2025_mTAW.xlsx': ('zeebrugge', 2025),
    }
    
    # Look for files in current dir and xlsx-getijtabellen-taw-2025 subdirectory
    search_paths = ['.', 'xlsx-getijtabellen-taw-2025']
    
    found_stations = {}
    for filename in stations.keys():
        for path in search_paths:
            full_path = os.path.join(path, filename)
            if os.path.exists(full_path):
                found_stations[full_path] = stations[filename]
                print(f"Found: {full_path}")
                break
        else:
            print(f"Warning: {filename} not found in {search_paths}")
    
    converted_files = []
    
    for excel_file, (station_name, year) in found_stations.items():
        try:
            json_file = convert_excel_to_json(excel_file, station_name, year)
            if json_file:
                converted_files.append(json_file)
        except Exception as e:
            print(f"Failed to convert {excel_file}: {e}")
            continue
    
    print(f"\nConversion complete. Created {len(converted_files)} JSON files:")
    for file in converted_files:
        print(f"  - {file}")
    
    if converted_files:
        print(f"\nTo use in the app, copy these files to:")
        print(f"  - App bundle (add to Xcode project), OR")
        print(f"  - App Documents directory for side-loading")

if __name__ == "__main__":
    main()
