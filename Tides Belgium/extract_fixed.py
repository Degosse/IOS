#!/usr/bin/env python3
"""Fixed extraction of tide data from Excel files"""

import openpyxl
import json
from datetime import datetime, timedelta

def parse_time(time_val):
    if time_val is None:
        return None
    if hasattr(time_val, 'hour'):
        return f'{time_val.hour:02d}:{time_val.minute:02d}'
    time_str = str(time_val).strip()
    if ':' in time_str:
        parts = time_str.split(':')
        if len(parts) >= 2:
            try:
                hour = int(parts[0])
                minute = int(parts[1])
                return f'{hour:02d}:{minute:02d}'
            except ValueError:
                pass
    return None

def parse_height(height_val):
    if height_val is None:
        return None
    try:
        if isinstance(height_val, (int, float)):
            return float(height_val)
        height_str = str(height_val).replace(',', '.')
        return float(height_str)
    except ValueError:
        return None

def get_months_from_sheet_name(sheet_name):
    """Convert Dutch sheet names to month numbers"""
    month_map = {
        'jan-feb': [1, 2],
        'mrt-apr': [3, 4], 
        'mei-jun': [5, 6],
        'jul-aug': [7, 8],
        'sept-okt': [9, 10],
        'nov-dec': [11, 12]
    }
    return month_map.get(sheet_name.lower(), [])

def extract_station_data(excel_path, station_name, year=2025):
    """Extract tide data from Excel file"""
    print(f"Processing {station_name} for {year}...")
    
    wb = openpyxl.load_workbook(excel_path)
    all_tides = []
    
    for sheet_name in wb.sheetnames:
        print(f"  Processing sheet: {sheet_name}")
        ws = wb[sheet_name]
        months = get_months_from_sheet_name(sheet_name)
        
        if not months:
            print(f"    Skipping unknown sheet: {sheet_name}")
            continue
            
        # Start from row 4 (after headers)
        for row in range(4, ws.max_row + 1):
            day_val = ws.cell(row=row, column=1).value
            
            # Skip header rows and invalid days
            if not isinstance(day_val, (int, float)) or day_val <= 0 or day_val > 31:
                continue
                
            day = int(day_val)
            
            # Process each possible month for this sheet
            for month in months:
                try:
                    # Validate date exists
                    test_date = datetime(year, month, day)
                    date_str = f'{year}-{month:02d}-{day:02d}'
                    
                    # Extract up to 4 tide times/heights from the row
                    tide_columns = [(3, 4), (5, 6), (10, 11)]  # Time-Height pairs
                    
                    for time_col, height_col in tide_columns:
                        time_val = ws.cell(row=row, column=time_col).value
                        height_val = ws.cell(row=row, column=height_col).value
                        
                        time_str = parse_time(time_val)
                        height = parse_height(height_val)
                        
                        if time_str and height is not None:
                            tide_type = 'high' if height >= 2.5 else 'low'  # Adjusted threshold
                            all_tides.append({
                                'date': date_str,
                                'time': time_str,
                                'height': round(height, 2),
                                'type': tide_type
                            })
                    
                    # Check continuation row
                    next_row = row + 1
                    if next_row <= ws.max_row:
                        next_day_val = ws.cell(row=next_row, column=1).value
                        # If next row doesn't have a day number, it's a continuation
                        if not isinstance(next_day_val, (int, float)):
                            for time_col, height_col in tide_columns:
                                time_val = ws.cell(row=next_row, column=time_col).value
                                height_val = ws.cell(row=next_row, column=height_col).value
                                
                                time_str = parse_time(time_val)
                                height = parse_height(height_val)
                                
                                if time_str and height is not None:
                                    tide_type = 'high' if height >= 2.5 else 'low'
                                    all_tides.append({
                                        'date': date_str,
                                        'time': time_str,
                                        'height': round(height, 2),
                                        'type': tide_type
                                    })
                    
                    break  # Successfully processed this day, move to next
                    
                except ValueError:
                    # Invalid date (e.g., Feb 30), try next month
                    continue
    
    # Sort and remove duplicates
    all_tides.sort(key=lambda x: (x['date'], x['time']))
    
    # Remove duplicates
    unique_tides = []
    seen = set()
    for tide in all_tides:
        key = f"{tide['date']}_{tide['time']}"
        if key not in seen:
            seen.add(key)
            unique_tides.append(tide)
    
    return unique_tides

# Process just Nieuwpoort first to test
try:
    tides = extract_station_data('xlsx-getijtabellen-taw-2025/Nieuwpoort2025_mTAW.xlsx', 'nieuwpoort')
    
    with open('nieuwpoort_2025_fixed.json', 'w') as f:
        json.dump(tides, f, indent=2)
    
    print(f"✓ Created nieuwpoort_2025_fixed.json with {len(tides)} tides")
    
    # Show August data specifically
    aug_tides = [t for t in tides if '2025-08-' in t['date']]
    print(f"August tides found: {len(aug_tides)}")
    for tide in aug_tides:
        if tide['date'] in ['2025-08-11', '2025-08-12', '2025-08-13']:
            print(f"  {tide['date']} {tide['time']}: {tide['height']}m ({tide['type']})")
    
except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()
